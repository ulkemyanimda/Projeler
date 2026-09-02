# -*- coding: utf-8 -*-
"""
SINIF & ÖĞRETMEN PLANLAMA SİSTEMİ
=================================

Girdi 1 (öğretmen dilekçe listesi) beklenen kolonlar:
    Adı Soyadı, Branş, SALI, ÇARŞAMBA, PERŞEMBE,
    Cumartesi 11, Cumartesi 13, Cumartesi 17, Cumartesi 18, Cumartesi 19,
    Pazar 11, Pazar 13, Pazar 17, Pazar 18, Pazar 19,
    2 saatten fazla ders almak istiyor mu?
    (X/x = o slotta derse girmek istiyor)

Girdi 2 (topluluk/ülke istatistiği) beklenen kolonlar:
    ToplulukID, Ülke Kategorisi, Öğrenci Sayısı, Yüzde
    ToplulukID formatı: <yas><seviye_no>  ör: cinar1, fidan3, filiz2, tohum1
        1=Başlangıç 2=Temel 3=Orta 4=İleri

Bu script tamamen GENEL amaçlıdır: aynı kolon adlarına sahip başka
dosyalar verildiğinde de (öğrenci/öğretmen sayısı değişse bile) çalışır.
"""

import re
import math
import argparse
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# 1) SABİTLER
# --------------------------------------------------------------------------

WEEKDAY_SLOTS = ['SALI', 'ÇARŞAMBA', 'PERŞEMBE']  # hepsi TR saati 18:00 kabul edilir
WEEKEND_HOURS = [11, 13, 17, 18, 19]
WEEKEND_DAYS = ['Cumartesi', 'Pazar']
WEEKEND_SLOTS = [f'{d} {h}' for d in WEEKEND_DAYS for h in WEEKEND_HOURS]
ALL_SLOTS = WEEKDAY_SLOTS + WEEKEND_SLOTS

# Ülkeye göre hangi slotlar uygun (TR saatine göre gece/gündüz kısıtı)
# ABD: yalnızca hafta sonu akşam (17/18/19) TR saati -> onlar için gündüz/akşam
# Çin: yalnızca hafta sonu sabah/öğlen (11/13) TR saati -> onlar için öğlen/akşam
# Diğer her şey (Avrupa, Diğer, vs.): tüm slotlar uygun (esnek havuz)
ABD_SLOTS = [f'{d} {h}' for d in WEEKEND_DAYS for h in (17, 18, 19)]
CIN_SLOTS = [f'{d} {h}' for d in WEEKEND_DAYS for h in (11, 13)]

MIN_CLASS, MAX_CLASS, TARGET_CLASS = 5, 15, 10
MAX_CLASSES_PER_TEACHER = 3  # bir öğretmene en fazla bu kadar sınıf verilir

GENEL_BRANSLAR_FALLBACK_ORDER = [
    'Türkçe', 'Türk Dili ve Edebiyatı', 'Sosyal Bilgiler', 'Fen Bilimleri',
    'İlköğretim Matematik', 'Din Kültürü ve Ahlak Bilgisi', 'Görsel Sanatlar',
    'Müzik', 'Bilgisayar ve Öğretim Teknolojileri', 'Almanca', 'Fizik',
    'Rehber Öğretmeni', 'Sınıf Öğretmeni', 'İngilizce', 'Okul Öncesi',
]

YAS_PREFIX_MAP = {'cinar': 'Çınar', 'çınar': 'Çınar', 'fidan': 'Fidan',
                   'filiz': 'Filiz', 'tohum': 'Tohum'}
SEVIYE_NO_MAP = {'1': 'Başlangıç', '2': 'Temel', '3': 'Orta', '4': 'İleri'}


def normalize(s):
    if s is None:
        return ''
    return str(s).strip()


def norm_key(s):
    return normalize(s).lower().replace('İ', 'i').replace('I', 'ı')


# --------------------------------------------------------------------------
# 2) SLOT ÇAKIŞMA MANTIĞI
#    Dersler 1.5 saat sürdüğü için aynı gün ardışık saat etiketleri
#    (17-18, 18-19) çakışır; 17-19 çakışmaz. 11-13 çakışmaz (2 saat ara).
#    Hafta içi günler birbirinden ve hafta sonundan tamamen bağımsızdır.
# --------------------------------------------------------------------------

def slot_day(slot):
    if slot in WEEKDAY_SLOTS:
        return slot
    return slot.split()[0]


def slot_hour(slot):
    m = re.search(r'(\d+)$', slot)
    return int(m.group(1)) if m else None


def slots_overlap(s1, s2):
    if s1 == s2:
        return True
    d1, d2 = slot_day(s1), slot_day(s2)
    if d1 != d2:
        return False
    if d1 in WEEKDAY_SLOTS:
        return False
    h1, h2 = slot_hour(s1), slot_hour(s2)
    if h1 is None or h2 is None:
        return False
    return abs(h1 - h2) == 1


# --------------------------------------------------------------------------
# 3) TOPLULUK ID -> YAŞ / SEVİYE ÇÖZÜMLEME
# --------------------------------------------------------------------------

def parse_topluluk_id(topluluk_id):
    m = re.match(r'([a-zçğıöşüA-ZÇĞİÖŞÜ]+)\s*(\d+)', normalize(topluluk_id))
    if not m:
        return None, None
    yas = YAS_PREFIX_MAP.get(m.group(1).lower())
    seviye = SEVIYE_NO_MAP.get(m.group(2))
    return yas, seviye


def teacher_priority_list(yas, seviye):
    """Bu yaş/seviye kombinasyonu için tercih edilecek branş sırası."""
    if yas == 'Tohum':
        order = ['Okul Öncesi']
        if seviye == 'Başlangıç':
            order.append('İngilizce')
    elif yas == 'Filiz':
        order = ['Sınıf Öğretmeni']
        if seviye == 'Başlangıç':
            order.append('İngilizce')
    else:  # Fidan, Çınar
        order = ['İngilizce'] if seviye == 'Başlangıç' else []
    order += [b for b in GENEL_BRANSLAR_FALLBACK_ORDER if b not in order]
    return order


# --------------------------------------------------------------------------
# 4) SINIF (GRUP) OLUŞTURMA — ülkeye göre AYRIŞTIRMADAN, zaman kısıtına göre
#    uygun havuzlara dağıtıp mümkün olan en dengeli / karma sınıfları kurar.
# --------------------------------------------------------------------------

def split_sizes(n, min_c=MIN_CLASS, max_c=MAX_CLASS, target=TARGET_CLASS):
    """n öğrenciyi min-max aralığında, ~target hedefli sınıflara böler.
    Dönen liste: her sınıfın öğrenci sayısı."""
    if n <= 0:
        return []
    if n <= max_c:
        return [n]
    k = max(1, round(n / target))
    while n / k > max_c:
        k += 1
    while k > 1 and n / k < min_c:
        k -= 1
    base = n // k
    rem = n % k
    sizes = [base + 1 if i < rem else base for i in range(k)]
    return sizes


def build_classes_for_topluluk(topluluk_id, country_counts):
    """
    country_counts: {'ABD': n, 'Çin': n, 'Avrupa': n, ...}
    Döndürür: liste of dict {sınıf_no, boyut, ülke_dagilimi, uygun_slotlar, uyari}
    Mantık:
      - ABD ve Çin öğrencileri kendi zaman havuzlarına (disjoint) kilitlidir.
      - Esnek havuz (Avrupa/Diğer) her iki tip sınıfa da eklenip onları
        5-15 aralığına tamamlamak / 10'a yaklaştırmak için kullanılır.
      - Kalan esnek öğrenciler kendi (tamamen serbest zamanlı) sınıflarını oluşturur.
      - Bu şekilde ülkelere göre KATI bir ayrım yapılmaz; sadece zaman
        kısıtı olan öğrenciler için zorunlu uygun slot seti belirlenir.
    """
    abd_n = country_counts.get('ABD', 0)
    cin_n = country_counts.get('Çin', 0)
    flex_n = sum(v for k, v in country_counts.items() if k not in ('ABD', 'Çin'))

    classes = []
    warnings = []

    def make_pool_classes(pool_name, n, eligible_slots):
        nonlocal flex_n
        sizes = split_sizes(n)
        pool_classes = []
        for size in sizes:
            used_flex = 0
            # 5'in altındaysa esnek öğrencilerle tamamla
            if size < MIN_CLASS and flex_n > 0:
                need = MIN_CLASS - size
                take = min(need, flex_n)
                flex_n -= take
                used_flex += take
                size += take
            # hedefe (10) yaklaştırmak için fırsatçı biçimde esnek ekle (opsiyonel, kapasiteyi 15'e kadar kullan)
            if size < TARGET_CLASS and flex_n > 0:
                top_up = min(TARGET_CLASS - size, flex_n, MAX_CLASS - size)
                if top_up > 0:
                    flex_n -= top_up
                    used_flex += top_up
                    size += top_up
            country_mix = {}
            base_country = 'ABD' if pool_name == 'ABD' else ('Çin' if pool_name == 'Çin' else 'Esnek/Avrupa')
            orig = size - used_flex
            if orig > 0:
                country_mix[base_country] = orig
            if used_flex > 0:
                country_mix['Esnek (Avrupa/Diğer)'] = country_mix.get('Esnek (Avrupa/Diğer)', 0) + used_flex
            w = None
            if size < MIN_CLASS:
                w = (f'UYARI: {topluluk_id} - {pool_name} havuzunda {size} kişilik sınıf '
                     f'minimum {MIN_CLASS} sınırının altında kaldı (esnek öğrenci yetersiz). '
                     f'Manuel birleştirme / başka toplulukla harmanlama gerekebilir.')
                warnings.append(w)
            pool_classes.append({
                'boyut': size,
                'ulke_dagilimi': country_mix,
                'uygun_slotlar': eligible_slots,
                'uyari': w,
            })
        return pool_classes

    if abd_n > 0:
        classes += make_pool_classes('ABD', abd_n, ABD_SLOTS)
    if cin_n > 0:
        classes += make_pool_classes('Çin', cin_n, CIN_SLOTS)
    if flex_n > 0:
        classes += make_pool_classes('Esnek', flex_n, ALL_SLOTS)

    for i, c in enumerate(classes, 1):
        c['sinif_no'] = i
        c['topluluk_id'] = topluluk_id

    return classes, warnings


# --------------------------------------------------------------------------
# 5) ZAMAN SLOTU ATAMA (sınıflar arası dengeli dağıtım, round-robin)
# --------------------------------------------------------------------------

def assign_time_slots(all_classes):
    """Her sınıfa, kendi uygun_slotlar listesinden round-robin biçimde
    (aynı toplulukta art arda aynı slotu tekrar etmemeye çalışarak) bir slot atar.
    Amaç: talebi slotlara yayıp öğretmen çakışmasını azaltmak."""
    cursor = defaultdict(int)  # uygun_slotlar tuple -> son kullanılan index
    for c in all_classes:
        key = tuple(c['uygun_slotlar'])
        idx = cursor[key] % len(key)
        c['zaman_dilimi'] = key[idx]
        cursor[key] += 1


# --------------------------------------------------------------------------
# 6) ÖĞRETMEN ATAMA
# --------------------------------------------------------------------------

def load_teachers(path):
    df = pd.read_excel(path)
    df.columns = [normalize(c) for c in df.columns]
    teachers = []
    for _, row in df.iterrows():
        name = normalize(row.get('Adı Soyadı'))
        if not name:
            continue
        brans = normalize(row.get('Branş'))
        available = set()
        for slot in ALL_SLOTS:
            col = slot if slot in WEEKDAY_SLOTS else slot  # aynı isimlendirme
            val = row.get(col)
            if normalize(val).upper() == 'X':
                available.add(slot)
        cok_ders = normalize(row.get('2 saatten fazla ders almak istiyor mu?'))
        wants_multi = cok_ders.strip().lower() == 'evet'
        teachers.append({
            'ad_soyad': name,
            'brans': brans,
            'brans_key': norm_key(brans),
            'musait_sloti': available,
            'coklu_ders_istiyor': wants_multi,
            'atanan_sloti': set(),      # bu koşuya özel, o slotu kullandı mı
            'atanan_sinif_sayisi': 0,
        })
    return teachers


def assign_teachers(all_classes, teachers):
    # Kıtlığı azaltmak için: önce özel branş gerektiren (Okul Öncesi, Sınıf Öğretmeni,
    # İngilizce başlangıç) sınıfları, sonra genel/esnek sınıfları ata.
    def scarcity_key(c):
        yas, seviye = parse_topluluk_id(c['topluluk_id'])
        pri = teacher_priority_list(yas, seviye)
        # ilk tercih branşındaki öğretmen sayısı ne kadar azsa o kadar önce işlensin
        first_pref = pri[0] if pri else None
        n_avail = sum(1 for t in teachers if t['brans_key'] == norm_key(first_pref))
        return n_avail

    all_classes_sorted = sorted(all_classes, key=scarcity_key)

    for c in all_classes_sorted:
        yas, seviye = parse_topluluk_id(c['topluluk_id'])
        pref_order = teacher_priority_list(yas, seviye)
        slot = c['zaman_dilimi']
        assigned = None
        used_fallback_any_brans = False

        def teacher_usable(t):
            if slot not in t['musait_sloti']:
                return False
            if t['atanan_sinif_sayisi'] >= MAX_CLASSES_PER_TEACHER:
                return False
            if t['atanan_sinif_sayisi'] == 0:
                return True
            if not t['coklu_ders_istiyor']:
                return False
            # birden fazla ders istiyor: çakışan bir slotu var mı kontrol et
            for used in t['atanan_sloti']:
                if slots_overlap(used, slot):
                    return False
            return True

        for brans in pref_order:
            candidates = [t for t in teachers
                          if t['brans_key'] == norm_key(brans) and teacher_usable(t)]
            if candidates:
                # en az yüklü / hiç atanmamış öğretmeni tercih et
                candidates.sort(key=lambda t: t['atanan_sinif_sayisi'])
                assigned = candidates[0]
                break

        if assigned is None:
            # inisiyatif: branş fark etmeksizin uygun herhangi bir öğretmen
            candidates = [t for t in teachers if teacher_usable(t)]
            if candidates:
                candidates.sort(key=lambda t: t['atanan_sinif_sayisi'])
                assigned = candidates[0]
                used_fallback_any_brans = True

        if assigned is not None:
            assigned['atanan_sloti'].add(slot)
            assigned['atanan_sinif_sayisi'] += 1
            c['ogretmen'] = assigned['ad_soyad']
            c['ogretmen_brans'] = assigned['brans']
            note = c.get('uyari')
            if used_fallback_any_brans:
                extra = (f'Not: {slot} için tercih edilen branştan uygun öğretmen bulunamadı, '
                         f'inisiyatif ile {assigned["brans"]} branşından atama yapıldı.')
                note = (note + ' | ' + extra) if note else extra
            c['uyari'] = note
        else:
            c['ogretmen'] = None
            c['ogretmen_brans'] = None
            w = f'ATANAMADI: {slot} slotunda uygun/müsait öğretmen bulunamadı.'
            c['uyari'] = (c['uyari'] + ' | ' + w) if c.get('uyari') else w

    return all_classes


# --------------------------------------------------------------------------
# 7) ANA AKIŞ
# --------------------------------------------------------------------------

def run(teachers_path, stats_path, output_path):
    df_stats = pd.read_excel(stats_path)
    df_stats.columns = [normalize(c) for c in df_stats.columns]
    df_stats = df_stats.dropna(subset=['ToplulukID'])

    all_classes = []
    all_warnings = []

    for topluluk_id, grp in df_stats.groupby('ToplulukID'):
        country_counts = {}
        for _, r in grp.iterrows():
            ulke = normalize(r['Ülke Kategorisi'])
            adet = int(r['Öğrenci Sayısı']) if not pd.isna(r['Öğrenci Sayısı']) else 0
            country_counts[ulke] = country_counts.get(ulke, 0) + adet
        classes, warns = build_classes_for_topluluk(topluluk_id, country_counts)
        all_classes += classes
        all_warnings += warns

    assign_time_slots(all_classes)

    teachers = load_teachers(teachers_path)
    assign_teachers(all_classes, teachers)

    write_output(all_classes, teachers, all_warnings, output_path)
    return all_classes, teachers, all_warnings


# --------------------------------------------------------------------------
# 8) EXCEL ÇIKTISI
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF')
WARN_FILL = PatternFill('solid', fgColor='FFF2CC')
UNASSIGNED_FILL = PatternFill('solid', fgColor='F8CBAD')
BASE_FONT = Font(name='Arial')


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'


def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_output(all_classes, teachers, warnings, output_path):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Sınıf Planı ----
    ws = wb.active
    ws.title = 'Sınıf Planı'
    headers = ['ToplulukID', 'Sınıf No', 'Öğrenci Sayısı', 'Ülke Dağılımı',
               'Zaman Dilimi', 'Atanan Öğretmen', 'Öğretmen Branşı', 'Not / Uyarı']
    ws.append(headers)
    all_classes_sorted = sorted(all_classes, key=lambda c: (c['topluluk_id'], c['sinif_no']))
    for c in all_classes_sorted:
        ulke_str = ', '.join(f'{k}: {v}' for k, v in c['ulke_dagilimi'].items())
        row = [
            c['topluluk_id'], c['sinif_no'], c['boyut'], ulke_str,
            c['zaman_dilimi'], c.get('ogretmen') or '(ATANMADI)',
            c.get('ogretmen_brans') or '-', c.get('uyari') or '',
        ]
        ws.append(row)
        r = ws.max_row
        for cell in ws[r]:
            cell.font = BASE_FONT
        if not c.get('ogretmen'):
            for cell in ws[r]:
                cell.fill = UNASSIGNED_FILL
        elif c.get('uyari'):
            for cell in ws[r]:
                cell.fill = WARN_FILL
    style_header(ws, len(headers))
    autofit(ws, [12, 9, 14, 30, 16, 24, 26, 60])

    # ---- Sheet 2: Öğretmen Yükü ----
    ws2 = wb.create_sheet('Öğretmen Yükü')
    headers2 = ['Adı Soyadı', 'Branş', 'Atanan Sınıf Sayısı', 'Atanan Zaman Dilimleri',
                'Çoklu Ders İstiyor mu?', 'Müsait Olduğu Slot Sayısı']
    ws2.append(headers2)
    for t in sorted(teachers, key=lambda x: -x['atanan_sinif_sayisi']):
        ws2.append([
            t['ad_soyad'], t['brans'], t['atanan_sinif_sayisi'],
            ', '.join(sorted(t['atanan_sloti'])),
            'Evet' if t['coklu_ders_istiyor'] else 'Hayır',
            len(t['musait_sloti']),
        ])
        for cell in ws2[ws2.max_row]:
            cell.font = BASE_FONT
    style_header(ws2, len(headers2))
    autofit(ws2, [26, 26, 18, 40, 18, 18])

    # ---- Sheet 3: Uyarılar / Manuel Kontrol ----
    ws3 = wb.create_sheet('Uyarılar')
    ws3.append(['Uyarı'])
    unique_warn_rows = 0
    for c in all_classes_sorted:
        if c.get('uyari'):
            ws3.append([f"[{c['topluluk_id']} - Sınıf {c['sinif_no']}] {c['uyari']}"])
            unique_warn_rows += 1
            for cell in ws3[ws3.max_row]:
                cell.font = BASE_FONT
                cell.fill = WARN_FILL
    if unique_warn_rows == 0:
        ws3.append(['Herhangi bir uyarı oluşmadı.'])
        ws3['A2'].font = BASE_FONT
    style_header(ws3, 1)
    autofit(ws3, [110])

    # ---- Sheet 4: Özet ----
    ws4 = wb.create_sheet('Özet')
    ws4.append(['ToplulukID', 'Toplam Öğrenci', 'Toplam Sınıf', 'Ortalama Sınıf Boyutu',
                'Atanamayan Sınıf Sayısı'])
    by_topluluk = defaultdict(list)
    for c in all_classes:
        by_topluluk[c['topluluk_id']].append(c)
    for tid, classes in sorted(by_topluluk.items()):
        total_students = sum(c['boyut'] for c in classes)
        n_classes = len(classes)
        avg = total_students / n_classes if n_classes else 0
        unassigned = sum(1 for c in classes if not c.get('ogretmen'))
        ws4.append([tid, total_students, n_classes, round(avg, 1), unassigned])
        for cell in ws4[ws4.max_row]:
            cell.font = BASE_FONT
    style_header(ws4, 5)
    autofit(ws4, [14, 16, 14, 20, 22])

    wb.save(output_path)


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Sınıf ve öğretmen planlama sistemi')
    parser.add_argument('--ogretmenler', required=True, help='Öğretmen dilekçe listesi .xlsx yolu')
    parser.add_argument('--istatistik', required=True, help='Topluluk/ülke istatistik .xlsx yolu')
    parser.add_argument('--cikti', default='sinif_plani.xlsx', help='Çıktı dosyası yolu')
    args = parser.parse_args()
    run(args.ogretmenler, args.istatistik, args.cikti)
    print(f'Tamamlandı -> {args.cikti}')


# python .\class_scheduler.py --ogretmenler ogretmen_dilekce_listesi.xlsx --istatistik abd_cin_diger_istatistik.xlsx
